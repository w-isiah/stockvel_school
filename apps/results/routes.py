from flask import (
    Blueprint, g, render_template, request, redirect,
    url_for, flash, session, jsonify, send_file, current_app
)
from werkzeug.utils import secure_filename
from datetime import datetime
from jinja2 import TemplateNotFound
from io import BytesIO

import os
import random
import re
import logging

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

import mysql.connector
from mysql.connector import Error

from apps.results import blueprint
from apps import get_db_connection

from openpyxl.styles import Font, Alignment




# Access the upload folder from the current Flask app configuration
def allowed_file(filename):
    """Check if the uploaded file has a valid extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']





@blueprint.route('/streams_data')
def streams_data():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute('''
        SELECT 
            s.stream_id,
            s.stream_name,
            c.class_id,
            c.class_name
        FROM stream s
        JOIN classes c ON s.class_id = c.class_id
        ORDER BY s.stream_name
    ''')
    streams = cursor.fetchall()

    # No need to manually convert if using dictionary cursor
    return jsonify(streams)
















@blueprint.route('/pdownload_template', methods=['GET'])
def pdownload_template():
    # Extract parameters
    class_id = request.args.get('class_id')
    year_id = request.args.get('year_id')
    term_id = request.args.get('term_id')
    subject_id = request.args.get('subject_id')
    assessment_id = request.args.get('assessment_id')
    stream_id = request.args.get('stream_id')

    if not all([class_id, year_id, term_id, subject_id, assessment_id, stream_id]):
        flash("Please select all required fields including stream.", "error")
        return redirect(url_for('results_blueprint.pupload_excel'))

    conn = get_db_connection()
    if not conn:
        flash("Failed to connect to the database.", "error")
        return redirect(url_for('results_blueprint.pupload_excel'))

    cursor = conn.cursor(dictionary=True)

    try:
        def fetch_column(query, column):
            cursor.execute(query)
            return [row[column] for row in cursor.fetchall()]

        # Reference data
        classes = fetch_column("SELECT class_name FROM classes", "class_name")
        study_years = fetch_column("SELECT year_name FROM study_year", "year_name")
        assessments = fetch_column("SELECT assessment_name FROM assessment", "assessment_name")
        terms = fetch_column("SELECT term_name FROM terms", "term_name")
        subjects = fetch_column("SELECT subject_name FROM subjects", "subject_name")
        streams = fetch_column("SELECT stream_name FROM stream", "stream_name")

        # Lookups
        def get_value(query, value):
            cursor.execute(query, (value,))
            row = cursor.fetchone()
            return row and list(row.values())[0]

        class_name = get_value("SELECT class_name FROM classes WHERE class_id = %s", class_id)
        stream_name = get_value("SELECT stream_name FROM stream WHERE stream_id = %s", stream_id)
        term_name = get_value("SELECT term_name FROM terms WHERE term_id = %s", term_id)
        assessment_name = get_value("SELECT assessment_name FROM assessment WHERE assessment_id = %s", assessment_id)
        year_name = get_value("SELECT year_name FROM study_year WHERE year_id = %s", year_id)
        subject_name = get_value("SELECT subject_name FROM subjects WHERE subject_id = %s", subject_id)

        if not all([class_name, stream_name, term_name, assessment_name, year_name, subject_name]):
            flash("Invalid selection detected. Please try again.", "error")
            return redirect(url_for('results_blueprint.pupload_excel'))

        # Pupils
        cursor.execute("""
            SELECT reg_no, first_name, other_name, last_name
            FROM pupils
            WHERE class_id = %s AND stream_id = %s AND year_id = %s AND term_id = %s
            ORDER BY last_name
        """, (class_id, stream_id, year_id, term_id))
        pupils = cursor.fetchall()

    finally:
        cursor.close()
        conn.close()

    if not pupils:
        flash("No pupils found for the selected filters.", "error")
        return redirect(url_for('results_blueprint.pupload_excel'))

    # Create Excel workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Results Template"

    # Title rows
    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')
    ws.merge_cells('A3:L3')

    ws['A1'] = "SACRED HEART PRIMARY SCHOOL - KYAMUSANSALA"
    ws['A2'] = "P.O. Box 1759, Masaka | Tel: 0772-848153 / 0702-253560"
    ws['A3'] = "Email: shps-rscjp@gmail.com"

    for row in range(1, 4):
        cell = ws[f'A{row}']
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')

    # Header row (row 4)
    headers = [
        "reg_no", "first_name", "other_name", "last_name",
        "class", "stream", "term", "assessment",
        "study_year", "notes", "subject", "mark"
    ]
    ws.append(headers)

    # Data rows (start from row 5)
    for pupil in pupils:
        ws.append([
            pupil['reg_no'], pupil['first_name'], pupil['other_name'], pupil['last_name'],
            class_name, stream_name, term_name, assessment_name, year_name, "", subject_name, ""
        ])

    # Reference sheet
    ref = wb.create_sheet("drop_down_data")
    ref.append(["classes", "streams", "study_years", "assessments", "terms", "subjects"])
    ref.append([
        ", ".join(classes),
        ", ".join(streams),
        ", ".join(study_years),
        ", ".join(assessments),
        ", ".join(terms),
        ", ".join(subjects)
    ])

    # Safe list string builder (max 255 characters)
    def safe_join(values):
        return ",".join([str(v).replace('"', "'") for v in values])[:255]

    validations = [
        (DataValidation(type="list", formula1=f'"{safe_join(classes)}"'), 'E'),
        (DataValidation(type="list", formula1=f'"{safe_join(streams)}"'), 'F'),
        (DataValidation(type="list", formula1=f'"{safe_join(terms)}"'), 'G'),
        (DataValidation(type="list", formula1=f'"{safe_join(assessments)}"'), 'H'),
        (DataValidation(type="list", formula1=f'"{safe_join(study_years)}"'), 'I'),
        (DataValidation(type="list", formula1=f'"{safe_join(subjects)}"'), 'K'),
    ]

    for dv, col in validations:
        ws.add_data_validation(dv)
        dv.add(f"{col}5:{col}{len(pupils) + 4}")

    # Send file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="shpsk_Marks_Upload_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )







@blueprint.route('/pupload_excel', methods=['GET', 'POST'])
def pupload_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch dropdown data
    cursor.execute('SELECT year_id, year_name AS study_year FROM study_year ORDER BY year_name')
    study_years = cursor.fetchall()

    cursor.execute('SELECT class_id, class_name FROM classes ORDER BY class_name')
    class_list = cursor.fetchall()

    cursor.execute('SELECT term_id, term_name FROM terms ORDER BY term_name')
    terms = cursor.fetchall()

    cursor.execute('SELECT subject_id, subject_name FROM subjects ORDER BY subject_name')
    subjects = cursor.fetchall()

    cursor.execute('SELECT assessment_id, assessment_name FROM assessment ORDER BY assessment_name')
    assessments = cursor.fetchall()

    cursor.execute('SELECT stream_id, stream_name FROM stream ORDER BY stream_name')
    streams = cursor.fetchall()

    if request.method == 'POST':
        file = request.files.get('file')

        # Validate file
        if not file or file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash('Invalid file format. Please upload an Excel (.xlsx) file.', 'danger')
            return redirect(request.url)

        # Save the uploaded file
        filename = secure_filename(file.filename)
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            import pandas as pd
            df = pd.read_excel(file_path)

            # Custom function to validate and extract the data
            processed_data, errors, existing_reg_nos, duplicate_reg_nos = validate_excel_data(df)

            if errors:
                flash('Errors encountered:\n' + '\n'.join(errors), 'danger')
                return redirect(request.url)

            if duplicate_reg_nos:
                flash(f"Duplicate reg_no(s) found: {', '.join(duplicate_reg_nos)}", 'danger')
                return redirect(request.url)

            if existing_reg_nos:
                flash(f"Existing reg_no(s): {', '.join(existing_reg_nos)} (skipped).", 'warning')

            if processed_data:
                insert_scores_into_database(processed_data)
                flash(f"{len(processed_data)} score record(s) uploaded successfully!", 'success')
            else:
                flash('No new records to insert.', 'info')

            return redirect(url_for('results_blueprint.pupload_excel'))

        except pd.errors.EmptyDataError:
            flash('Uploaded file is empty.', 'danger')
        except Exception as e:
            flash(f'Error processing the file: {str(e)}', 'danger')

        return redirect(url_for('results_blueprint.pupload_excel'))

    # GET request: render form with all dropdowns
    return render_template(
        'results/upload_excel.html',
        study_years=study_years,
        class_list=class_list,
        terms=terms,
        assessments=assessments,
        subjects=subjects,
        streams=streams  # Make sure the template uses this
    )














def validate_excel_data(df):
    processed_data = []
    errors = []
    existing_reg_nos = []
    seen_keys = set()

    # Normalize column names
    df.columns = df.columns.str.lower()

    # Required columns
    required_columns = {'reg_no', 'class', 'stream', 'study_year', 'term', 'assessment', 'subject', 'mark'}
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Get user ID from session
    user_id = session.get('id')
    if not user_id:
        raise ValueError("Missing or invalid user session ID.")

    try:
        with get_db_connection() as connection:
            cursor = connection.cursor()

            # Mapping queries
            mapping_queries = {
                'classes': "SELECT class_name, class_id FROM classes",
                'study_year': "SELECT year_name, year_id FROM study_year",
                'terms': "SELECT term_name, term_id FROM terms",
                'assessment': "SELECT assessment_name, assessment_id FROM assessment",
                'subjects': "SELECT subject_name, subject_id FROM subjects",
                'streams': "SELECT stream_name, stream_id FROM stream"
            }

            mappings = {}
            for key, query in mapping_queries.items():
                cursor.execute(query)
                mappings[key] = {row[0].strip(): row[1] for row in cursor.fetchall()}

            # Load existing score combinations to avoid duplicates
            cursor.execute("""
                SELECT reg_no, class_id, stream_id, year_id, term_id, assessment_id, subject_id 
                FROM scores
            """)
            existing_score_keys = {
                (row[0], row[1], row[2], row[3], row[4], row[5], row[6]) for row in cursor.fetchall()
            }

    except Exception as e:
        errors.append(f"Database error: {str(e)}")
        return processed_data, errors, existing_reg_nos, []

    # Iterate over Excel rows
    for index, row in df.iterrows():
        if row.isnull().all():
            continue

        reg_no = str(row.get('reg_no')).strip()
        class_name = str(row.get('class')).strip()
        stream_name = str(row.get('stream')).strip()
        year_name = str(row.get('study_year')).strip()
        term_name = str(row.get('term')).strip()
        assessment_name = str(row.get('assessment')).strip()
        subject_name = str(row.get('subject')).strip()
        mark = row.get('mark')
        notes = row.get('notes') if 'notes' in row else None

        # Check for missing critical fields
        if not all([reg_no, class_name, stream_name, year_name, term_name, assessment_name, subject_name]):
            errors.append(f"Row {index + 2}: Missing one or more required fields.")
            continue

        if pd.isna(mark):
            errors.append(f"Row {index + 2}: 'Mark' is missing.")
            continue

        # Map names to IDs
        class_id = mappings['classes'].get(class_name)
        stream_id = mappings['streams'].get(stream_name)
        year_id = mappings['study_year'].get(year_name)
        term_id = mappings['terms'].get(term_name)
        assessment_id = mappings['assessment'].get(assessment_name)
        subject_id = mappings['subjects'].get(subject_name)

        # Collect errors for missing mappings
        row_errors = []
        if not class_id:
            row_errors.append(f"Class '{class_name}' not found.")
        if not stream_id:
            row_errors.append(f"Stream '{stream_name}' not found.")
        if not year_id:
            row_errors.append(f"Study year '{year_name}' not found.")
        if not term_id:
            row_errors.append(f"Term '{term_name}' not found.")
        if not assessment_id:
            row_errors.append(f"Assessment '{assessment_name}' not found.")
        if not subject_id:
            row_errors.append(f"Subject '{subject_name}' not found.")

        if row_errors:
            errors.append(f"Row {index + 2}: " + "; ".join(row_errors))
            continue

        # Check for duplicates
        key = (reg_no, class_id, stream_id, year_id, term_id, assessment_id, subject_id)
        if key in existing_score_keys:
            existing_reg_nos.append(reg_no)
            continue

        # Append validated row
        data = {
            'user_id': user_id,
            'reg_no': reg_no,
            'class_id': class_id,
            'stream_id': stream_id,
            'year_id': year_id,
            'term_id': term_id,
            'assessment_id': assessment_id,
            'subject_id': subject_id,
            'mark': mark,
            'notes': None if pd.isna(notes) else notes
        }

        processed_data.append(data)

    return processed_data, errors, existing_reg_nos, []







def insert_scores_into_database(processed_data):
    if not processed_data:
        print("⚠️ No score data to insert.")
        return

    if not isinstance(processed_data, list) or not all(isinstance(item, dict) for item in processed_data):
        raise ValueError("❌ processed_data must be a list of dictionaries.")

    for idx, data in enumerate(processed_data):
        if 'mark' not in data or pd.isna(data['mark']):
            raise ValueError(f"❌ 'mark' is missing or null in record at index {idx}: {data}")
        data['notes'] = None if pd.isna(data.get('notes')) else data.get('notes')

    insert_query = """
        INSERT INTO scores (
            user_id, reg_no, class_id, stream_id, term_id, year_id,
            assessment_id, subject_id, mark, notes
        ) VALUES (
            %(user_id)s, %(reg_no)s, %(class_id)s, %(stream_id)s, %(term_id)s, %(year_id)s,
            %(assessment_id)s, %(subject_id)s, %(mark)s, %(notes)s
        )
    """

    check_existing_query = """
        SELECT COUNT(*) FROM scores
        WHERE reg_no = %(reg_no)s
        AND class_id = %(class_id)s
        AND stream_id = %(stream_id)s
        AND year_id = %(year_id)s
        AND term_id = %(term_id)s
        AND assessment_id = %(assessment_id)s
        AND subject_id = %(subject_id)s
    """

    try:
        with get_db_connection() as connection:
            with connection.cursor() as cursor:
                inserted_count = 0

                for data in processed_data:
                    cursor.execute(check_existing_query, {
                        'reg_no': data['reg_no'],
                        'class_id': data['class_id'],
                        'stream_id': data['stream_id'],
                        'year_id': data['year_id'],
                        'term_id': data['term_id'],
                        'assessment_id': data['assessment_id'],
                        'subject_id': data['subject_id']
                    })
                    existing_count = cursor.fetchone()[0]

                    if existing_count == 0:
                        cursor.execute(insert_query, data)
                        inserted_count += 1

                connection.commit()
                print(f"✅ Successfully inserted {inserted_count} record(s).")
    except Exception as e:
        print(f"❌ Error inserting data: {e}")
        raise











  
@blueprint.route('/delete_result/<int:results_id>')
def delete_result(results_id):
    """Deletes a results from the database."""
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Delete the results with the specified ID
        cursor.execute('DELETE FROM results WHERE result_id = %s', (results_id,))
        connection.commit()
        flash("results deleted successfully.", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('results_blueprint.results'))




@blueprint.route('/<template>')
def route_template(template):

    try:

        if not template.endswith('.html'):
            template += '.html'

        # Detect the current page
        segment = get_segment(request)

        # Serve the file (if exists) from app/templates/home/FILE.html
        return render_template("results/" + template, segment=segment)

    except TemplateNotFound:
        return render_template('home/page-404.html'), 404

    except:
        return render_template('home/page-500.html'), 500


# Helper - Extract current page name from request
def get_segment(request):

    try:

        segment = request.path.split('/')[-1]

        if segment == '':
            segment = 'results'

        return segment

    except:
        return None
